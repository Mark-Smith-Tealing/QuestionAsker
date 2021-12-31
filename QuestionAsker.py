
#ud# -*- coding: utf-8 -*-
"""
Created on Sun May 24 13:01:51 2020

@author: mark
"""

import tkinter as tk
import xlrd
import random
import openpyxl as py
import os
import sympy as sp
import matplotlib
import matplotlib.pyplot as plt

from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfilename
from IPython.display import display_latex
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

matplotlib.use('TkAgg')

class QuestionAskerApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)                    # Initialise Tk
        tk.Tk.wm_title(self, "Question Asker")  # Set the title of the window to Question Asker
        
        window = tk.Frame(self)                 # Create a frame
        window.grid(row=0,column=0)             # Place frame into grid
        window.grid_rowconfigure(0, weight=1)   # Configure row
        window.grid_columnconfigure(0, weight=1)# Configure Column
        
        QuestionAskerApp.QuestionToBeAsked = tk.StringVar()     # String to hold Question
        QuestionAskerApp.AnswerForAskedQuestion = tk.StringVar()# String to hold real answer
        QuestionAskerApp.TheirAnswer = tk.StringVar()           # String to hold their answer
        QuestionAskerApp.CorrectAnswers = tk.IntVar()           # Int to hold number of Correct answers
        QuestionAskerApp.TotalAnswers = tk.IntVar()             # Int to hold number questions asked

        self.frames = {}    # Create Dictionary to hold each window
        
        for F in (Menu, Study, Question, Answer, Tryme, Trymetext, Recap):
            
            frame = F(window, self)                         # Create a frame for each window
            self.frames[F] = frame                          # Add each frame into the dictionary 
            frame.grid(row=0, column = 0, sticky = "nsew")  # Place each frame into position
        
        self.show_frame(Menu)   # Show the first window - menu
        
    def show_frame(self, cont):     # Define the function to show a frame
        frame = self.frames[cont]   # Set active frame to the desired frame 
        frame.tkraise()             # Raise the desired frame to the front
        
        if cont == Question:                                                                                                    # If we are rasing the Question frame
            Question.QuestionActual1.config(text=QuestionAskerApp.QuestionToBeAsked.get())                                      # Update the Question text label
            Question.label1.config(text = 'Questions Answered Correctly: %s / %s' % (QuestionAskerApp.CorrectAnswers.get(),     # Update the Questions taken label
                                                                                     QuestionAskerApp.TotalAnswers.get()))      # 
            Question.QuestionSheet.config(text = Study.SheetName)   # Update the sheet name
        
        elif cont == Answer:                                                                                    # If we are raising the Answer frame
            Answer.TheirAnswerActual.config(text=QuestionAskerApp.TheirAnswer.get())                            # Update their Answer text label
            Answer.QuestionActual.config(text = QuestionAskerApp.QuestionToBeAsked.get(), bg = "light blue")    # Update the Question text label
            Answer.SheetAnswerLabel.config(text = Study.SheetName)                                              # Update the Study sheet name
            
            cont.ax.clear()                                                                        # Clear the figure
            cont.ax.set_title('The Answer')                                                        # Set the title of the figure
            cont.ax.text(0.05,0.8, QuestionAskerApp.AnswerForAskedQuestion.get(), fontsize = 10)   # Set the text in the figure
            cont.canvas.draw()                                                                     # Draw the canvas
        
        elif cont == Menu:                                          # If we are raising the Menu
            
            if Menu.count != 0:                                     # ... And it is not the first instance of the menu 
                Menu.openFileEntry.config(text=Menu.new_file_name)  # Set the file name to the selected file
        
        elif cont == Trymetext:                                         # If we are raising the Trymetext
            Trymetext.ax.clear()                                        # Clear the figure
            Trymetext.ax.text(0.05,0.8, Tryme.temptext, fontsize = 10)  # Set the text in the figure
            Trymetext.canvas.draw()                                     # Draw the canvas
        
        elif cont == Recap:                                                                     # If we are raising the Recap page
            a =  QuestionAskerApp.CorrectAnswers.get()                                          # Set a to be the number of correct answers
            b = QuestionAskerApp.TotalAnswers.get()                                             # Set b to the number of questions asked
            
            if b != 0:                                           # If we get at least one correct...
                c = a/b*100                                      # ...Set c to the percentage
            
            var = '%'                                                                      # Variable for a percentage sign to add to text
            cont.Totalquestionsansweredlabel2.config(text = a)                             # Update the questions asked
            cont.Totalquestionsaskedlabel2.config(text = b)                                # Update the correct answers
            cont.Totalquestionspercentagelabel.config(text = '%s %s' % (round(c), var))    # Update the percentage label 
            cont.progressBar.config(value=QuestionAskerApp.CorrectAnswers.get(), maximum= QuestionAskerApp.TotalAnswers.get())
        
class Menu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        self.controller = controller
        
        Menu.count = 0              # Count for the number of times that the menu has been shown
        global currentWeek          # Int current week
        global questionsDone        # Int number of questions done
        global questionsToBeDone    # Int number of questions to be answered for the relevant week
    
        currentWeek = 0             # Set current week to 0
        questionsToBeDone = 184     # Set questions to be done to 184
        questionsDone = 170         # Set Questions done to 170
    
        logo = tk.Text(self, width=8, height=3) # Leave space for a logo
        logo.grid(row=0,column=0,stick="w")     # Place the logo at (0,0) in grid
    
        menuTitle = ttk.Label(self, text="QuestionAsker")   # Label to give title of page
        menuTitle.grid(row=0,column=1)                      # Place the title label at top at (0,1)
    
        studyButton = ttk.Button(self, 
                                 text="Study", 
                                 command=lambda: controller.show_frame(Study))  # Create a study button
        studyButton.grid(row=1,column=1)                                        # place the button at (1,1)
    
        statsButton = ttk.Button(self, 
                                 text="Try Me Text",
                                 command=lambda: controller.show_frame(Tryme)) # Create a stats button
        statsButton.grid(row=2,column=1)             # Place the button at (2,1)
    
        optionsButton = ttk.Button(self, text="Options") # Create an options button
        optionsButton.grid(row=3,column=1)               # Place the button at (3,1)
    
        progressLabel = ttk.Label(self, text = 'Progress: Week %s' % currentWeek)   # Create a progress label
        progressLabel.grid(row=4, columnspan = 1, sticky = 'w')                     # Place the label at (4,1)
    
        progressNumbers = ttk.Label(self, text = ' %s / %s Completed' % (questionsDone, questionsToBeDone)) # Create a progress label
        progressNumbers.grid(row=5, columnspan = 1, sticky = 'w')                                           # Place the label at (5,1)
    
        changeWeekLabel = ttk.Label(self, text = "Change Week") # Create a change week label
        changeWeekLabel.grid(row=5, column=1, columnspan=2)     # Place the label at (5,1)
    
        progressBar = ttk.Progressbar(self, orient = 'horizontal',  # Create a weekly progress bar
                                      length = 150,                 # Set the length to 150
                                       value = questionsDone,       # Set tje value to the number of questions done
                                      phase = 1,                    # Set the  phase to 1
                                      maximum = questionsToBeDone,  # Set the max value to the questions to be done
                                      mode = 'determinate')         # Set the mode to determinate
        progressBar.grid(row=6,column=0)                            # Place the progress bar at (6,0)
    
        changeWeekButtonDown = ttk.Button(self, text = "<<")    # Create a change week down button
        changeWeekButtonUp = ttk.Button(self, text = ">>")      # Create a change week up button
        changeWeekButtonDown.grid(row=6,column=1)               # Place the change week down button at (6,1)
        changeWeekButtonUp.grid(row=6,column=2)                 # Place the change week up button at (6,2)
        
        openFile = ttk.Label(self, text = "\nLoaded File Name:")    # Create an open file label
        openFile.grid(row=7,column=0, sticky = "w")                 # Place the label at (7,0)
        
        Menu.new_file_name = ""                                 # Create a placeholder file name
        Menu.file_name = os.getcwd()                            # Get the current working directory
        Menu.first_name = Menu.file_name + "\Example Questions" # Set the first file name to the appropriate wd
        
        Menu.openFileEntry = ttk.Label(self, text = Menu.first_name, width=70)  # Create a label with the file name
        Menu.openFileEntry.grid(row=8,column=0, columnspan=4)                   # Place the label at (8,0)
        
        changeFileButton = ttk.Button(self, text = "Change File", command = lambda : change_file()) # Create a button to change file
        changeFileButton.grid(row=9, column=0, sticky="w")                                          # Place the button at (9,0)
        
        def change_file():                          # Define a function to change file
            Menu.count += 1                         # Add to the menu count
            Menu.new_file_name = askopenfilename()  # Ask for a new file
            controller.show_frame(Menu)             # Show a certain frame

class Study(tk.Frame):
    def __init__(self, parent, controller):         
        self.controller = controller                
        tk.Frame.__init__(self,parent)
        
        label=tk.Label(self, text = "Study")        # Create a study title
        label.grid(row=0,column=1,sticky="nswe")    # Place the title at (0,1)
    
        logo = tk.Text(self, width=8, height=3) # Create a space for the logo
        logo.grid(row=0,column=0,sticky="w")    # Place the logo at (0,0)
        
        studyLabel = tk.Label(self, text="Select Sheets to Study")  # Create a label to select sheets
        studyLabel.grid(row=1,column=0, sticky="w")                 # Place the label at (1,0)
        
        for index, sheet_name in enumerate(sheets.keys()):  # Run through the sheets
            SelectedSheetsBool[sheet_name] = tk.IntVar()    # Fill the SelectedSheetsBool dictionary with Intvars for the buttons
            ProgressBars[sheet_name] = tk.IntVar()          # Fill the ProgressBars dictionary with Intvars for the progress bars
            
            check = ttk.Checkbutton(self, 
                                    text = sheet_name, 
                                    variable = SelectedSheetsBool[sheet_name])              # Create a check button for each sheet with variable in dictionary           
            progressBar = ttk.Progressbar(self, 
                                      orient = 'horizontal', 
                                      length = 150, 
                                      value = questionsDone, 
                                      phase = 1, 
                                      maximum = questionsToBeDone, 
                                      mode = 'determinate')                                 # Create a progress bar for each sheet
            progressLabel = ttk.Label(self, 
                                      text='%s / %s' % (questionsDone, questionsToBeDone))  # Create a progress label for each progress bar
            
            progressBar.grid(row=index+2, column=1)     # Place the progress bar at (index+2,1)
            check.grid(row=index+2, sticky='w')         # Place the checkbutton at (index+2,0)
            progressLabel.grid(row=index+2, column = 2) # Place the progress label at (index+2,2)
        
        def Finish(event=None):                         # Finish Function runs when we exit page
            for i in SelectedSheetsBool.keys():         # For the sheets...
                if SelectedSheetsBool[i].get() == 1:    # If the button is selected...
                   SelectedSheets[i] = sheets[i]        # ... add the relevant sheet to the list of active sheets
                
                elif SelectedSheetsBool[i].get() == 0:  # Elseif the button isn't selected...
                    if i in SelectedSheets:             # ... and if that sheet is one that is selected...
                        del SelectedSheets[i]           # ... remove that sheet from the selected sheets
            
            if "p" in SelectedSheets:                   # If the placeholder remains...
                del SelectedSheets["p"]                 # Delete the placeholder
            
            Study.SheetName = random.choice(list(SelectedSheets.keys()))                            # Make a random choice from the sheet names
            Study.ActiveSheet = (SelectedSheets[Study.SheetName])                                   # Set the active sheet to that randomly picked sheet name
            Study.QuestionNumber = random.randint(0,len(Study.ActiveSheet)-1)                       # Make a random choice from the questions in that sheet
            
            QuestionAskerApp.QuestionToBeAsked.set(Study.ActiveSheet[Study.QuestionNumber][0])      # Set the question to the new random choice
            QuestionAskerApp.AnswerForAskedQuestion.set(Study.ActiveSheet[Study.QuestionNumber][1]) # Set the answer to the new random choice
            
            controller.show_frame(Question)                                                         # Show the Question frame
        
        studyButton = ttk.Button(self, text="Go!", command=Finish)  # Create the go button
        studyButton.grid(row=len(sheet_names)+4,column=2)           # Place the button at (len(sheet_names)+4,2)
        
        progressLabel = tk.Label(self, text = '\nProgress: Week %s' % currentWeek)  # Create a progress label
        progressLabel.grid(row = len(sheet_names)+2, columnspan = 2, sticky = 'w')  # Place the label in the grid
    
        progressNumbers = tk.Label(self, text = ' %s / %s Completed' % (questionsDone, questionsToBeDone))  # Create a progress label
        progressNumbers.grid(row=len(sheet_names)+3, columnspan = 2, sticky = 'w')                          # Place the label in the grid 
    
        progressBar = ttk.Progressbar(self, 
                                      orient = 'horizontal', 
                                      length = 150, 
                                      value = questionsDone, 
                                      phase = 1, 
                                      maximum = questionsToBeDone, 
                                      mode = 'determinate')                 # Create a progress bar to show the number of questions done and to be done
        progressBar.grid(row=len(sheet_names)+4, columnspan=2, sticky="w")  # Place the progress bar in the grid
        
class Question(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        self.controller = controller
        
        label=ttk.Label(self, text = "Question:")    # Create the title label
        label.grid(row=0,sticky="nswe")              # Place the title label in the grid
        labelsheet= ttk.Label(self, text = "Sheet:") # Create a sheet label
        labelsheet.grid(row=1,sticky="nswe")         # Place the sheet label in the grid
        
        Study.SheetName = "" # Create a placeholder for the sheetname
        Question.QuestionActual1 = tk.Label(self, 
                                            text = QuestionAskerApp.QuestionToBeAsked.get(),
                                            bg = "light blue")                                   # Create the label to ask the question
        Question.QuestionSheet = ttk.Label(self, text = Study.SheetName)
        TheirAnswerActual = tk.Text(self)                                                        # Create a text box for user to input data             
        
        Question.QuestionActual1.grid(row=0)    # Place the question label in the grid
        Question.QuestionSheet.grid(row=1)      # Place the sheet label in the grid
        TheirAnswerActual.grid(row=2)           # Place the text box in the grid
    
        def Finish(event=None):                                                     # Finish function to leave question page
            QuestionAskerApp.TheirAnswer.set(TheirAnswerActual.get("1.0", "end"))   # Get the user's answer from the textbox
            TheirAnswerActual.delete("1.0", "end")                                  # Delete the user's answer
            
            controller.show_frame(Answer)                                           # Show the answer frame
            QuestionAskerApp.TotalAnswers.set(QuestionAskerApp.TotalAnswers.get()+1)# Update the number of questions asked 

        ttk.Button(self, text = "Okay!", command = Finish).grid(columnspan=2)       # Create an okay button 
        Question.label1 = tk.Label(self, 
                                   text = 'Questions Answered Correctly: %s / %s' % (QuestionAskerApp.CorrectAnswers.get(), 
                                                                                     QuestionAskerApp.TotalAnswers.get()))  # Create a label giving number of questions asked correctly
        Question.label1.grid(row=4)                                                 # Place the label in the grid
     
class Answer(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        self.controller = controller
        
        label=tk.Label(self, text = "Answer")   # Create the answer title
        label.grid(row=0,column=0,sticky="nswe")# Place the answer title in the grid
        
        QuestionAskerApp.CorrectAnswers.set(0)  # Set the number of correct answers to zero
        
        Answer.QuestionLabel = tk.Label(self, text = "Question: ")                          # Create the Question label 
        Answer.QuestionActual = tk.Label(self, 
                                         text = QuestionAskerApp.QuestionToBeAsked.get(), 
                                         bg = "light blue")                                 # Create the Question asked label
        
        Answer.SheetLabel = tk.Label(self, text = "Sheet:")                                 # Create a sheet label
        Answer.SheetAnswerLabel = tk.Label(self, text = Study.SheetName)                    # Create a sheet name label
    
        Answer.TheirAnswerLabel = tk.Label(self, text = "Your Answer was: ")                # Create their Answer label
        Answer.TheirAnswerActual = tk.Label(self, text = QuestionAskerApp.TheirAnswer.get())# Create their Answer label with the users typed answer
        
        graphFrame = tk.Frame(self) # Create a frame for the graph
        
        fig = Figure(figsize=(8,3), dpi = 100) # Create a figure
        Answer.ax = fig.add_subplot(111)       # Add a subplot to the figure

        Answer.canvas = FigureCanvasTkAgg(fig, graphFrame)   # Create a canvas
        Answer.canvas.get_tk_widget().pack()                 # Pack the canvas
        Answer.canvas._tkcanvas.pack()                       # pack the canvas

        Answer.ax.get_xaxis().set_visible(False) # Remove the x axis
        Answer.ax.get_yaxis().set_visible(False) # Remove the y axis
        
        toolbar = NavigationToolbar2Tk(Answer.canvas, graphFrame)  # Create a toolbar
        toolbar.update()                                           # Update the toolbar
        Answer.canvas._tkcanvas.pack()            # Pack the toolbar
        
        graphFrame.grid(row=2, columnspan=4)    # Place the graph frame in the grid
        
        Answer.QuestionLabel.grid(row=0)                # Place the labels in the grid
        Answer.QuestionActual.grid(row=0,column=1) 
        Answer.SheetLabel.grid(row=1)
        Answer.SheetAnswerLabel.grid(row=1,column=1)
        Answer.TheirAnswerLabel.grid(row=3)
        Answer.TheirAnswerActual.grid(row=3,column=1)
        
        CheckVarYes = tk.IntVar()   # Create a checkvar for the yes checkbutton
        CheckVarNo = tk.IntVar()    # Create a checkvar for the No checkbutton
        CheckVarMeh = tk.IntVar()   # Create a checkvar for the Meh checkbutton
        
        checkbutton_label = tk.Label(self, text= "\n\ndid you get it right?")               # Create a label to ask if the user got the question right
        checkbutton_widget1 = tk.Checkbutton(self, text = "Yes", variable = CheckVarYes)    # Create a checkbutton for Yes
        checkbutton_widget2 = tk.Checkbutton(self, text = "No", variable = CheckVarNo)      # Create a checkbutton for No
        checkbutton_widget3 = tk.Checkbutton(self, text = "Meh", variable = CheckVarMeh)    # Create a checkbutton for Meh
        
        checkbutton_label.grid(row=4)           # Place the label in the grid
        checkbutton_widget1.grid(row=5)         # Place the yes checkbutton in the grid
        checkbutton_widget2.grid(row=5,column=1)# Place the no checkbutton in the grid
        checkbutton_widget3.grid(row=5,column=2)# Place the meh checkbutton in the grid
        
        def Finish(event=None):     # Create a finish function for when we leave the Answer page
            
            global CorrectAnswers   # Creat the global variable correctanswers
            QuestionAskerApp.CorrectAnswers.set(QuestionAskerApp.CorrectAnswers.get()+CheckVarYes.get()) # Add the value of the yes checkbutton's checkvar to the correct answers 
            
            sheets[Study.SheetName][Study.QuestionNumber][2] += 1                   # Add one to the number of times this question has been asked 
            sheets[Study.SheetName][Study.QuestionNumber][3] += CheckVarYes.get()   # Add one or zero to the number of times this question has been answered correctly
            
            Study.SheetName = random.choice(list(SelectedSheets.keys()))                            # Pick a new random sheet
            Study.ActiveSheet = (SelectedSheets[Study.SheetName])                                   # Set the active sheet to the selected random sheet
            Study.QuestionNumber = random.randint(0,len(Study.ActiveSheet)-1)                       # Pick a random question from the sheet
            
            QuestionAskerApp.QuestionToBeAsked.set(Study.ActiveSheet[Study.QuestionNumber][0])      # Set the question to be asked to the selected question
            QuestionAskerApp.AnswerForAskedQuestion.set(Study.ActiveSheet[Study.QuestionNumber][1]) # Set the answer to the question to be asked
            Answer.QuestionActual.config(text=QuestionAskerApp.QuestionToBeAsked.get())             # Change the text on the label that asks a question
            
            CheckVarYes.set(0)              # Set the Yes button to off
            CheckVarNo.set(0)               # Set the No button to off
            CheckVarMeh.set(0)              # Set the Meh button to off
            controller.show_frame(Question) # Show the question frame 
            
        def BackToMainMenu(event=None):             # Define a function to take you back to the menu
            for keyName in (sheet_names):           # for each sheet...
                worksheet = xl_workbook[keyName]    # ... Set the active sheet to keyname
                
                for j in range(2,len(sheets[keyName])+2):                           # For each question...
                    worksheet.cell(row=j,column=3).value = sheets[keyName][j-2][2]  # Write the times asked to the correct cell
                    worksheet.cell(row=j,column=4).value = sheets[keyName][j-2][3]  # Write the times answered correctly to the correct cell
            
            xl_workbook.save(loadfilename)                              # Save the workbook
            
            for i in SelectedSheetsBool.keys(): # For each sheet...
                SelectedSheetsBool[i].set(0)    # Set the selected sheets all to zero
            controller.show_frame(Recap)        # Show the menu
            
        nextQuestion = ttk.Button(self, 
                                  text = "Next Question!", 
                                  command = Finish)         # Create the next question button
        nextQuestion.grid(row=6,column = 1)                 # Place the next question button in the grid
        
        button2= ttk.Button(self, 
                            text="Finish!", 
                            command=BackToMainMenu) # Create a finish button
        button2.grid(row=8,column=1)                # Place the button in the grid

class Tryme(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        self.controller = controller
        
        tryMeTitle=ttk.Label(self, text = "Try Me Text Page", font = "12")    # Create the title label

        informationLabel = tk.Label(self, 
                                   text = "Type your text into the box, press okay and it will be converted into latex")  # Create the label to give some information
        Tryme.ConvertibleText = tk.Text(self)                                                                             # Create a text box for user to input data             
        
        informationLabel.grid(row=1)        # Place the inormation label in the grid
        tryMeTitle.grid(row=0)              # Place the title label in the grid
        Tryme.ConvertibleText.grid(row=2)   # Place the text box in the grid
        Tryme.temptext = ""                 # Set a placeholder for the textbox
    
        def Finish(event=None):                                         # Finish function to leave question page
            Tryme.temptext = Tryme.ConvertibleText.get("1.0", "end")    # Get the text from the textbox
            controller.show_frame(Trymetext)                            # Show the trymetext frame            
            Tryme.ConvertibleText.delete("1.0", "end")                  # Delete the user's answer

        def ReturnToMenu(event=None):
            Tryme.ConvertibleText.delete("1.0", "end")  # Delete the user's answer
            controller.show_frame(Menu)                 # Show the Menu

        ttk.Button(self, text = "Okay!", command = Finish).grid(columnspan=2)   # Create an okay button 
        ttk.Button(self, text = "Finish!", command = ReturnToMenu).grid(columnspan=2)   # Create an okay button 
        
class Trymetext(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        self.controller = controller
        
        tryMeTitle=ttk.Label(self, text = "Try Me Text Page")    # Create the title label
                
        tryMeTitle.grid(row=0, columnspan = 2)          # Place the title label in the grid
        
        graphFrame = tk.Frame(self) # Create a frame for the graph
        
        fig = Figure(figsize=(8,2), dpi = 100) # Create a figure
        Trymetext.ax = fig.add_subplot(111)     # Add a subplot to the figure

        Trymetext.canvas = FigureCanvasTkAgg(fig, graphFrame) # Create a canvas
        Trymetext.canvas.get_tk_widget().pack()                 # Pack the canvas
        Trymetext.canvas._tkcanvas.pack()                       # pack the canvas
        
        Trymetext.ax.get_xaxis().set_visible(False) # Remove the x axis
        Trymetext.ax.get_yaxis().set_visible(False) # Remove the y axis
        
        toolbar = NavigationToolbar2Tk(Trymetext.canvas, graphFrame)  # Create a toolbar
        toolbar.update()                                        # Update the toolbar
        Trymetext.canvas._tkcanvas.pack(expand=True)            # Pack the toolbar
        
        graphFrame.grid(row=1, columnspan=2)    # Place the frame in the grid
    
        def Finish(event=None):          # Finish function 
            controller.show_frame(Tryme) # Show the answer frame
            
        def ReturnToMenu(event=None):   # Back to Menu function
            controller.show_frame(Menu) # Show the Menu

        ttk.Button(self, text = "Again!", command = Finish).grid(row=2,column=0,sticky="e")          # Create a try again button 
        ttk.Button(self, text = "Finish!", command = ReturnToMenu).grid(row=2,column=1,sticky="w")   # Create a return to menu button 

class Recap(tk.Frame):
        def __init__(self, parent, controller):
            tk.Frame.__init__(self,parent)
            self.controller = controller
            
            tryMeTitle=ttk.Label(self, text = "Recap")    # Create the title label
                
            tryMeTitle.grid(row=0, columnspan = 2)          # Place the title label in the grid
            
            Totalquestionsaskedlabel = ttk.Label(self, text = "Total Qs: ")                                     # Label for the total qs
            Totalquestionsansweredlabel = ttk.Label(self, text = "Correct As: ")                                # Label for the Correct As
            Totalquestionspercentagelabel = ttk.Label(self, text = "Percentage: ")                              # Label for Percentage
            
            Recap.Totalquestionsansweredlabel2 = ttk.Label(self, text = QuestionAskerApp.CorrectAnswers.get())  # Label for the total qs
            Recap.Totalquestionsaskedlabel2 = ttk.Label(self, text = QuestionAskerApp.TotalAnswers.get())       # Label for the correct As
            Recap.Totalquestionspercentagelabel = ttk.Label(self, text = "")                                    # Label for the percentage
            
            Totalquestionsaskedlabel.grid(row=1)    # Place the labels in the grid
            Totalquestionsansweredlabel.grid(row=2)
            Totalquestionspercentagelabel.grid(row=3)
            
            Recap.Totalquestionsaskedlabel2.grid(row=1, column=1)
            Recap.Totalquestionsansweredlabel2.grid(row=2, column=1)
            Recap.Totalquestionspercentagelabel.grid(row=3, column=1)
            
            def ReturnToMenu(event=None):   # Back to Menu function
                QuestionAskerApp.CorrectAnswers.set(0)                                                  # Reset the questions answered correctly to 0
                QuestionAskerApp.TotalAnswers.set(0)                                                    # Reset the questions asked to 0
                controller.show_frame(Menu) # Show the Menu        
            
            Recap.progressBar = ttk.Progressbar(self, 
                                          orient = 'horizontal', 
                                          length = 150, 
                                          value = QuestionAskerApp.CorrectAnswers.get(), 
                                          phase = 1, 
                                          maximum = QuestionAskerApp.TotalAnswers.get(), 
                                          mode = 'determinate')                                 # Create a progress bar for each sheet
            
            Recap.progressBar.grid(row=4,column=1,sticky='w')
            ttk.Button(self, text = "Finish!", command = ReturnToMenu).grid(row=5,column=1,sticky="w")   # Create a return to menu button 
        
os.chdir('C:/Users/mark/Downloads')
loadfilename = 'Notes.xlsx'  
xl_workbook = py.load_workbook(loadfilename)    # Load the example questions workbook

global sheet_names                      # Create a global variable to hold the sheet names

sheet_names = xl_workbook.sheetnames    # Load the sheet names into the variable

global masterlist                       # Create a global list to hold the questions 

CorrectAnswers = 0                      # Set correct answers to zero

masterlist = []                         # Create an empty list
currentlist = []                        # Create an empty list

sheets = {}                         # Create an empty dictionary to hold all sheet information
SelectedSheetsBool = {}             # Create an empty dictionary to hold which sheets are active
ProgressBars = {}                   # Create an empty dictionary to hold the progress bars
SelectedSheets = {"p": [["",""]]}   # Create an empty dictionary to hold the selected sheets (with placeholder)

for index, i in enumerate(sheet_names):                 # For each sheet...
    worksheet = xl_workbook[i]                          # ... Set to the active sheet
    num_cols = worksheet.max_column                     # Number of columns
    for row_idx in range(2, worksheet.max_row+1):       # Iterate through rows
        for col_idx in range(1, num_cols+1):            # Iterate through columns
            cell_obj = worksheet.cell(row_idx, col_idx) # Get cell object by row, col
            currentlist.append(cell_obj.value)          # Append cell to the current list
        masterlist.append(currentlist)                  # Append current list to master list
        currentlist = []                                # Clear current list
    sheets[i] = masterlist                              # Add masterlist to the dictionary
    masterlist = []                                     # Clear masterlist

app = QuestionAskerApp()    # Create an instance of the QuestionAskerApp Class
app.mainloop()              # Run the mainloop